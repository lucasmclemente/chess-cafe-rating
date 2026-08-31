"""Le a planilha do Chess Cafe e recalcula o rating de todos, partida a partida.

Replica exatamente as formulas da planilha:
  esperado = 1 / (1 + 10^((rating_adversario - rating_jogador) / 400))
  variacao = ROUND(40 * (pontuacao - esperado), 0)
  o adversario recebe a variacao com o sinal invertido (jogo de soma zero)
"""
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook

K = 40
RATING_PADRAO = 1500
PONTOS = {"1-0": 1.0, "0,5-0,5": 0.5, "0.5-0.5": 0.5, "0-1": 0.0}


def arredondar(x):
    """ROUND do Excel: empate arredonda para longe do zero (o round() do Python nao faz isso)."""
    return int(Decimal(str(x)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _mapear_colunas(aba):
    """Localiza colunas pelo nome do cabecalho, para a planilha poder ganhar colunas novas."""
    mapa = {}
    for celula in next(aba.iter_rows(min_row=1, max_row=1)):
        if isinstance(celula.value, str) and celula.value.strip():
            mapa[celula.value.strip().lower()] = celula.column - 1
    return mapa


def ler_planilha(caminho):
    wb = load_workbook(caminho, data_only=True)

    if "RATING" not in wb.sheetnames or "PARTIDAS" not in wb.sheetnames:
        raise SystemExit(f"A planilha precisa ter as abas RATING e PARTIDAS. Encontrei: {wb.sheetnames}")

    aba_rating = wb["RATING"]
    col = _mapear_colunas(aba_rating)
    c_nome, c_inicial = col.get("jogador"), col.get("rating inicial")
    if c_nome is None:
        raise SystemExit("Nao achei a coluna 'Jogador' na aba RATING.")

    iniciais = {}
    for linha in aba_rating.iter_rows(min_row=2, values_only=True):
        nome = linha[c_nome]
        if not isinstance(nome, str) or not nome.strip():
            continue
        bruto = linha[c_inicial] if c_inicial is not None else None
        iniciais[nome.strip()] = int(bruto) if isinstance(bruto, (int, float)) else RATING_PADRAO

    aba_partidas = wb["PARTIDAS"]
    col = _mapear_colunas(aba_partidas)
    c_a, c_b, c_res = col.get("jogador a"), col.get("jogador b"), col.get("resultado")
    c_data = col.get("data")
    if None in (c_a, c_b, c_res):
        raise SystemExit("Na aba PARTIDAS preciso das colunas 'Jogador A', 'Jogador B' e 'Resultado'.")

    partidas, avisos = [], []
    for n_linha, linha in enumerate(aba_partidas.iter_rows(min_row=2, values_only=True), start=2):
        a = linha[c_a].strip() if isinstance(linha[c_a], str) else None
        b = linha[c_b].strip() if isinstance(linha[c_b], str) else None
        res = linha[c_res].strip() if isinstance(linha[c_res], str) else None
        if not a and not b and not res:
            continue
        if not (a and b and res):
            avisos.append(f"linha {n_linha}: partida incompleta, ignorada")
            continue
        if res not in PONTOS:
            avisos.append(f"linha {n_linha}: resultado '{res}' nao reconhecido, ignorada")
            continue
        data = linha[c_data] if c_data is not None and c_data < len(linha) else None
        partidas.append({"linha": n_linha, "a": a, "b": b, "resultado": res, "data": data})

    return iniciais, partidas, avisos


def calcular(iniciais, partidas):
    """Roda as partidas em ordem e devolve o estado de cada jogador com seu historico."""
    jogadores = {}

    def registrar(nome):
        if nome not in jogadores:
            jogadores[nome] = {
                "nome": nome,
                "inicial": iniciais.get(nome, RATING_PADRAO),
                "rating": iniciais.get(nome, RATING_PADRAO),
                "v": 0, "e": 0, "d": 0,
                "historico": [],
            }
        return jogadores[nome]

    for nome in iniciais:
        registrar(nome)

    for n, p in enumerate(partidas, start=1):
        ja, jb = registrar(p["a"]), registrar(p["b"])
        ra, rb = ja["rating"], jb["rating"]

        pontos_a = PONTOS[p["resultado"]]
        esperado_a = 1 / (1 + 10 ** ((rb - ra) / 400))
        delta = arredondar(K * (pontos_a - esperado_a))

        ja["rating"] += delta
        jb["rating"] -= delta

        for jog, adv, r_antes, d, pts in ((ja, jb, ra, delta, pontos_a), (jb, ja, rb, -delta, 1 - pontos_a)):
            jog["v" if pts == 1 else "d" if pts == 0 else "e"] += 1
            jog["historico"].append({
                "n": n,
                "adversario": adv["nome"],
                "rating_adversario": rb if jog is ja else ra,
                "pontos": pts,
                "delta": d,
                "rating": jog["rating"],
                "data": p["data"].strftime("%d/%m/%Y") if hasattr(p["data"], "strftime") else (str(p["data"]) if p["data"] else None),
            })

        p["rating_a"], p["rating_b"], p["delta"] = ra, rb, delta

    ordenados = sorted(jogadores.values(), key=lambda j: (-j["rating"], j["nome"]))
    posicao, anterior = 0, None
    for i, j in enumerate(ordenados, start=1):
        if j["rating"] != anterior:
            posicao, anterior = i, j["rating"]
        j["posicao"] = posicao
        j["partidas"] = len(j["historico"])
        j["variacao"] = j["rating"] - j["inicial"]
    return ordenados
